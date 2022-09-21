import os
import csv
import zipfile
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import codecs

def test_create_zip():
    # Create ZIP file
    file_zip = zipfile.ZipFile('./resources/test.zip', 'w')
    # added files to ZIP
    file_zip.write('./resources/TestOne.pdf')
    file_zip.write('./resources/TestTwo.xlsx')
    file_zip.write('./resources/TestThree.csv')


def test_file_pdf():
    with zipfile.ZipFile(os.path.abspath('./resources/test.zip')) as test_zip_pdf:
        with test_zip_pdf.open('resources/TestOne.pdf') as test_pdf:
            test_pdf = PdfReader(test_pdf)
            text_check = test_pdf.pages[0].extract_text()
            assert ('Foundation Level Syllabus' in text_check)


def test_file_xlsx():
    with zipfile.ZipFile(os.path.abspath('./resources/test.zip')) as test_zip_xlsx:
        with test_zip_xlsx.open('resources/TestTwo.xlsx') as test_xlsx:
            test_xlsx = load_workbook(test_xlsx)
            sheet = test_xlsx.active
            data = sheet.cell(row=29, column=1).value
            assert data == 'Тихонович Анастасия Павловна'


def test_file_csv():
    with zipfile.ZipFile(os.path.abspath('./resources/test.zip')) as test_zip_csv:
        with test_zip_csv.open('resources/TestThree.csv') as test_csv:
            test_csv = csv.reader(codecs.iterdecode(test_csv, 'utf-8'))
            for line_number, line in enumerate(test_csv, 1):
                if line_number == 5:
                    assert line[5] == 'M'